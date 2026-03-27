"""Read attendee data from Excel file"""

from openpyxl import load_workbook
from typing import List, Dict


def read_attendees(excel_path: str) -> List[Dict[str, str]]:
    """
    Read attendee data from Excel file.

    Extracts columns:
    - Column B: Full Name
    - Column E: Food Restrictions

    Returns list of dicts with keys: 'name', 'restrictions'
    """
    attendees = []

    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Iterate through rows (skip header at row 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Column B = index 1, Column E = index 4
        name_cell = ws[f'B{row_idx}']
        restrictions_cell = ws[f'E{row_idx}']

        name = name_cell.value
        restrictions = restrictions_cell.value

        # Skip empty rows
        if not name:
            continue

        # Handle None/empty restrictions
        if restrictions is None:
            restrictions = "None"

        attendees.append({
            'name': str(name).strip(),
            'restrictions': str(restrictions).strip()
        })

    wb.close()
    return attendees
