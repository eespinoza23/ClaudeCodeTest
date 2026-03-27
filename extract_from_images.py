"""
Extract attendee data from images using OCR and create Excel file
"""

import pytesseract
from PIL import Image
import re
from openpyxl import Workbook
from pathlib import Path
import os
import subprocess

# Configure pytesseract to find tesseract.exe
# Try the actual installed location first, then fallback locations
tesseract_paths = [
    r'C:\Users\learn\AppData\Local\Programs\Tesseract-OCR\tesseract.exe',
    r'C:\Program Files\Tesseract-OCR\tesseract.exe',
    r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
]

tesseract_exe = None
for path in tesseract_paths:
    if Path(path).exists():
        pytesseract.pytesseract.pytesseract_cmd = path
        tesseract_exe = path
        # Also add to PATH environment variable
        tesseract_dir = str(Path(path).parent)
        os.environ['PATH'] = tesseract_dir + os.pathsep + os.environ.get('PATH', '')
        print(f"Found Tesseract at: {path}")
        break
else:
    print("WARNING: Tesseract not found at any standard location")

def extract_text_from_images(image_paths):
    """Extract text from multiple images using OCR"""
    all_text = ""

    for image_path in image_paths:
        print(f"Processing: {image_path}")
        try:
            img = Image.open(image_path)
            # Try pytesseract first
            try:
                text = pytesseract.image_to_string(img, lang='eng')
            except Exception as ptes_err:
                # Fallback: use subprocess directly if tesseract_exe is available
                if tesseract_exe:
                    print(f"  pytesseract failed, trying subprocess directly...")
                    try:
                        result = subprocess.run(
                            [tesseract_exe, str(image_path), 'stdout', '-l', 'eng'],
                            capture_output=True,
                            text=True,
                            timeout=30
                        )
                        text = result.stdout
                        if result.returncode != 0:
                            print(f"  Tesseract stderr: {result.stderr}")
                    except Exception as sub_err:
                        print(f"  Subprocess also failed: {sub_err}")
                        text = ""
                else:
                    raise ptes_err

            all_text += text + "\n"
        except Exception as e:
            print(f"Error processing {image_path}: {e}")

    return all_text

def parse_attendee_data(text):
    """Parse OCR text to extract attendee records"""
    attendees = []
    seen_emails = set()

    lines = text.split('\n')
    # Email pattern that allows underscores/hyphens but requires proper domain structure
    email_pattern = r'([a-zA-Z0-9._-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'

    # First pass: extract all emails with their immediate context
    emails_found = {}  # email -> {line_idx, name, after_text}
    names_found = {}   # line_idx -> name

    for line_idx, line in enumerate(lines):
        line = line.strip()

        # Skip empty and header lines
        if not line or any(header in line.upper() for header in ['EMAIL', 'FORMAT', 'CONDITIONAL', 'MERGE', 'COPY', 'NARROW', 'SHEET', 'PASTE', 'READY', 'ACCESSIBILITY']):
            continue

        # Find emails in this line
        for email_match in re.finditer(email_pattern, line):
            email = email_match.group(1).lower()

            # Validate email has proper structure (no spaces in domain)
            if email not in seen_emails and ' ' not in email:
                seen_emails.add(email)
                before = line[:email_match.start()].strip()
                after = line[email_match.end():].strip()

                # Try to extract name from this line
                name = ""
                if after:
                    words = after.split()
                    for word in words:
                        if any(kw in word.lower() for kw in ['count', 'no', 'none', 'food', 'vegetarian', 'vegan', 'gluten',
                                                               'dairy', 'nut', 'squad', 'day', 'allergy', 'intolerant', 'halal', 'meat', 'pork', 'chicken', 'fish', 'shellfish']):
                            break
                        name += word + " "
                    name = name.strip()

                if not name and before and not any(c in before for c in ['%', '$', '&', '@', '<', '>', '=']):
                    words = before.split()
                    if words:
                        name = ' '.join(words[-3:])

                emails_found[email] = {
                    'line_idx': line_idx,
                    'name': name,
                    'after_text': after
                }

        # Extract standalone names (lines that look like names but have no email)
        if '@' not in line and len(line.split()) >= 1 and len(line.split()) <= 3:
            # This might be a standalone name
            if not any(kw in line.lower() for kw in ['format', 'count', 'squad', 'day', 'none', '%', 'sheet', 'ready']):
                name = re.sub(r'[^a-zA-Z\s\'-.]', '', line).strip()
                if name and len(name) >= 3 and len(name.split()) <= 3:
                    names_found[line_idx] = name

    # Second pass: match standalone names to nearby emails
    for email, email_info in emails_found.items():
        name = email_info['name']

        # If no name found on the email line, look for nearest name
        if not name or len(name) < 3:
            line_idx = email_info['line_idx']
            # Look for names within 5 lines before or after
            for offset in range(1, 6):
                # Check before
                if line_idx - offset in names_found:
                    candidate = names_found[line_idx - offset]
                    if len(candidate) >= 3:
                        name = candidate
                        break
                # Check after
                if line_idx + offset in names_found:
                    candidate = names_found[line_idx + offset]
                    if len(candidate) >= 3:
                        name = candidate
                        break

        # Sanitize name
        name = re.sub(r'[^a-zA-Z\s\'-.]', '', name).strip()
        if len(name.split()) > 4:
            name = ' '.join(name.split()[:3])

        # Parse restrictions
        restrictions = "None"
        if email_info['after_text']:
            after_lower = email_info['after_text'].lower()
            for restriction in ['vegetarian', 'vegan', 'gluten', 'dairy', 'nut', 'lactose', 'halal', 'pork', 'meat', 'chicken', 'fish', 'shellfish', 'allergy', 'carb']:
                if restriction in after_lower:
                    restrictions = restriction.capitalize()
                    break

        # Add attendee if we have a good name
        if name and len(name) >= 3 and len(name.split()) >= 1:
            attendees.append([
                email,
                name,
                "Count me in",
                "Count me in",
                restrictions,
                "Squad 1"
            ])

    return attendees

def create_excel_file(attendees, output_path):
    """Create Excel file from attendee data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendees"

    # Header
    ws['A1'] = "Email"
    ws['B1'] = "Full Name"
    ws['C1'] = "Day One"
    ws['D1'] = "Day Two"
    ws['E1'] = "Food Restrictions"
    ws['F1'] = "Squad"

    # Data
    for row_idx, attendee in enumerate(attendees, start=2):
        for col_idx, value in enumerate(attendee, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)
    print(f"Created {output_path} with {len(attendees)} attendees")

def main():
    print("=" * 50)
    print("Attendee Data Extractor (OCR)")
    print("=" * 50)

    # Image paths
    desktop = Path.home() / "Desktop"
    image_paths = [
        desktop / "1.jpeg",
        desktop / "2.jpeg",
        desktop / "3.jpeg",
        desktop / "4.jpeg",
    ]

    # Check if images exist
    existing_images = [p for p in image_paths if p.exists()]
    if not existing_images:
        print("ERROR: No images found on Desktop!")
        return

    print(f"Found {len(existing_images)} images")

    # Extract text
    print("\nExtracting text from images...")
    text = extract_text_from_images(existing_images)

    # Debug: save raw OCR text for inspection
    print(f"\nDEBUG: Extracted {len(text)} characters of text")
    with open("ocr_output.txt", "w", encoding="utf-8") as f:
        f.write(text)
    print("Raw OCR text saved to ocr_output.txt")

    if text.strip():
        print("First 500 characters of extracted text:")
        print("-" * 50)
        print(text[:500])
        print("-" * 50)

    # Parse data
    print("Parsing attendee data...")
    attendees = parse_attendee_data(text)

    print(f"\nExtracted {len(attendees)} attendees")

    if attendees:
        print("\nFirst 5 attendees:")
        for attendee in attendees[:5]:
            print(f"  {attendee[0]} - {attendee[1]}")

    # Create Excel
    output_path = Path.cwd() / "qr-generator" / "attendees_extracted.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\nCreating Excel file...")
    create_excel_file(attendees, output_path)

    print("\nDone!")
    print(f"File saved to: {output_path}")

if __name__ == "__main__":
    main()
